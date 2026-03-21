import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

df = pd.read_excel("planilha_vendas_baguncada.xlsx")

df = df.dropna(how="all")
df = df.drop_duplicates()

df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
df["Quantidade"] = pd.to_numeric(df["Quantidade"], errors="coerce")
df["Preço"] = pd.to_numeric(df["Preço"], errors="coerce")

df = df[
    (df["Quantidade"].notna()) &
    (df["Preço"].notna()) &
    (df["Quantidade"] > 0) &
    (df["Preço"] > 0)
]

df = df.dropna(subset=["Produto", "Categoria", "Data"])
df["Vendedor"] = df["Vendedor"].fillna("Desconhecido")

df["Total"] = df["Quantidade"] * df["Preço"]

faturamento_total = df["Total"].sum()

ranking_produtos = df.groupby("Produto")["Quantidade"].sum().sort_values(ascending=False)
ranking_vendedores = df.groupby("Vendedor")["Total"].sum().sort_values(ascending=False)

print(f"Total de registros finais: {len(df)}")
print(f"Faturamento total: R$ {faturamento_total:.2f}")

with pd.ExcelWriter("Relatorio_Vendas_Limpo.xlsx", engine="openpyxl") as writer:


    df.to_excel(writer, index=False, sheet_name="Sheet1")
    worksheet = writer.sheets["Sheet1"]

    cor_cabecalho = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    fonte_branca = Font(color="FFFFFF", bold=True)

    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = cor_cabecalho
        cell.font = fonte_branca
        cell.alignment = Alignment(horizontal="center")

    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[col_letter].width = max_length + 5
    
    worksheet.freeze_panes = "A2"

    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=1).number_format = "DD/MM/YYYY"

    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=5).number_format = "R$ #,##0.00"

    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=worksheet.max_column).number_format = "R$ #,##0.00"
    
    linha_inicial = len(df) + 3

    worksheet.cell(row=linha_inicial, column=1, value="Faturamento Total")
    worksheet.cell(row=linha_inicial, column=2, value=faturamento_total)

    produto_top_nome = ranking_produtos.index[0]
    produto_top_valor = ranking_produtos.iloc[0]
    worksheet.cell(row=linha_inicial + 2, column=1, value="Produto mais Vendido")
    worksheet.cell(row=linha_inicial + 2, column=2, value=produto_top_nome)
    worksheet.cell(row=linha_inicial + 2, column=3, value=produto_top_valor)

    vendedor_nome = ranking_vendedores.index[0]
    vendedor_valor = ranking_vendedores.iloc[0]
    worksheet.cell(row=linha_inicial + 4, column=1, value="Vendedor Top")
    worksheet.cell(row=linha_inicial + 4, column=2, value=vendedor_nome)
    worksheet.cell(row=linha_inicial + 4, column=3, value=vendedor_valor)

    worksheet.cell(row=linha_inicial, column=1).font = Font(bold=True)
    worksheet.cell(row=linha_inicial + 2, column=1).font = Font(bold=True)
    worksheet.cell(row=linha_inicial + 4, column=1).font = Font(bold=True)

    worksheet.cell(row=linha_inicial, column=2).number_format = "R$ #,##0.00"
    worksheet.cell(row=linha_inicial + 4, column=3).number_format = "R$ #,##0.00"

print("Relatório gerado com sucesso!")